import logging
import os
from datetime import datetime, timedelta
import openpyxl
from aiogram import Router, F
from aiogram.filters import Command
from aiogram.types import Message

from config import ADMIN_USERS, WEEKDAYS
from services import get_next_shift
from typing import Tuple, List

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

router = Router()


@router.message(Command("test_files"))
async def cmd_test_files(message: Message):
    """Проверка наличия и состояния файлов расписания"""
    if str(message.from_user.id) not in ADMIN_USERS:
        return

    try:
        debug_info = ["📁 Проверка файлов расписания:\n"]

        files_to_check = ['schedule.xlsx', 'schedule_next.xlsx']
        for file in files_to_check:
            if os.path.exists(file):
                file_size = os.path.getsize(file)
                modified_time = datetime.fromtimestamp(os.path.getmtime(file))
                debug_info.extend([
                    f"Файл {file}:",
                    f"✅ Существует",
                    f"📊 Размер: {file_size} байт",
                    f"🕒 Изменен: {modified_time.strftime('%d.%m.%Y %H:%M:%S')}\n"
                ])
            else:
                debug_info.extend([
                    f"Файл {file}:",
                    f"❌ Не найден\n"
                ])

        await message.answer("\n".join(debug_info))

    except Exception as e:
        await message.answer(f"🚫 Ошибка при проверке файлов: {str(e)}")


@router.message(Command("test_schedule_structure"))
async def cmd_test_schedule_structure(message: Message):
    """Проверка структуры файлов расписания"""
    if str(message.from_user.id) not in ADMIN_USERS:
        return

    try:
        debug_info = ["📊 Проверка структуры расписания:\n"]

        for file in ['schedule.xlsx', 'schedule_next.xlsx']:
            if not os.path.exists(file):
                debug_info.append(f"❌ Файл {file} не найден\n")
                continue

            wb = openpyxl.load_workbook(file)
            ws = wb.active

            debug_info.extend([
                f"Файл {file}:",
                f"📏 Количество строк: {ws.max_row}",
                f"📏 Количество колонок: {ws.max_column}",
                "\n👥 Проверка первых 5 строк:"
            ])

            # Проверяем первые 5 строк
            for row in range(1, min(6, ws.max_row + 1)):
                row_data = [str(ws.cell(row=row, column=col).value) for col in range(1, 5)]
                debug_info.append(f"Строка {row}: {' | '.join(row_data)}")

            debug_info.append("")
            wb.close()

        await message.answer("\n".join(debug_info))

    except Exception as e:
        await message.answer(f"🚫 Ошибка при проверке структуры: {str(e)}")


@router.message(Command("test_user_schedule"))
async def cmd_test_user_schedule(message: Message):
    """Проверка расписания конкретного пользователя"""
    if str(message.from_user.id) not in ADMIN_USERS:
        return

    try:
        # Можно добавить аргумент для проверки разных пользователей
        args = message.text.split()
        test_login = args[1] if len(args) > 1 else "sm_kirillts"

        debug_info = [f"👤 Проверка расписания для {test_login}:\n"]

        for file in ['schedule.xlsx', 'schedule_next.xlsx']:
            if not os.path.exists(file):
                continue

            wb = openpyxl.load_workbook(file)
            ws = wb.active

            debug_info.append(f"Файл {file}:")

            # Поиск пользователя
            found = False
            for row in ws.iter_rows():
                cell = row[2]  # 3-я колонка
                if cell.value and str(cell.value).lower() == test_login.lower():
                    found = True
                    debug_info.append(f"✅ Пользователь найден в строке {cell.row}")

                    # Показываем смены на ближайшие дни
                    shifts = []
                    for col in range(4, min(ws.max_column + 1, 10)):  # Первые 6 дней
                        shift = ws.cell(row=cell.row, column=col).value
                        shifts.append(f"День {col - 3}: {shift}")
                    debug_info.append("Смены:")
                    debug_info.extend(shifts)
                    break

            if not found:
                debug_info.append("❌ Пользователь не найден")

            debug_info.append("")
            wb.close()

        # Тестируем функцию получения смен
        debug_info.append("🔄 Тест функции get_next_shift:")
        shifts = await get_next_shift(test_login)
        debug_info.append(shifts)

        await message.answer("\n".join(debug_info))

    except Exception as e:
        await message.answer(f"🚫 Ошибка при проверке расписания: {str(e)}")


@router.message(Command("test_date_handling"))
async def cmd_test_date_handling(message: Message):
    """Проверка обработки дат и переходов между месяцами"""
    if str(message.from_user.id) not in ADMIN_USERS:
        return

    try:
        current_date = datetime.now()
        debug_info = ["📅 Проверка обработки дат:\n"]

        # Генерируем тестовые даты
        test_dates = [
            current_date + timedelta(days=1),
            current_date + timedelta(days=2),
            current_date + timedelta(days=3),
            current_date.replace(day=1) + timedelta(days=32),  # следующий месяц
        ]

        for date in test_dates:
            debug_info.extend([
                f"Дата: {date.strftime('%d.%m.%Y')}",
                f"День недели: {WEEKDAYS[date.strftime('%A').lower()]}",
                f"Месяц: {date.month} {'(следующий)' if date.month != current_date.month else '(текущий)'}",
                f"Номер колонки: {date.day + 3}\n"
            ])

        await message.answer("\n".join(debug_info))

    except Exception as e:
        await message.answer(f"🚫 Ошибка при проверке дат: {str(e)}")


async def test_shift_command(test_login: str = "sm_kirillts") -> Tuple[bool, str]:
    """Тестирование команды shift"""
    try:
        shifts = await get_next_shift(test_login)

        # Проверяем корректность ответа
        if not shifts:
            return False, "Пустой ответ"

        # Проверяем наличие дат и информации о сменах
        lines = shifts.split('\n')
        if len(lines) < 4:  # Заголовок + 3 дня
            return False, "Неполный ответ"

        # Проверяем формат каждой строки
        for line in lines[1:]:  # Пропускаем заголовок
            if not line:
                continue
            # Проверяем формат строки (дата (день недели) - смена)
            if not ('202' in line and '(' in line and ')' in line and '-' in line):
                return False, f"Неверный формат строки: {line}"

        return True, "OK"
    except Exception as e:
        return False, str(e)


async def test_files_existence() -> Tuple[bool, str]:
    """Тестирование наличия файлов"""
    try:
        files = ['schedule.xlsx', 'schedule_next.xlsx']
        missing = [f for f in files if not os.path.exists(f)]
        if missing:
            return False, f"Отсутствуют файлы: {', '.join(missing)}"
        return True, "OK"
    except Exception as e:
        return False, str(e)


async def test_schedule_structure() -> Tuple[bool, str]:
    """Тестирование структуры расписания"""
    try:
        for file in ['schedule.xlsx', 'schedule_next.xlsx']:
            if not os.path.exists(file):
                continue

            wb = openpyxl.load_workbook(file)
            ws = wb.active

            # Проверяем минимальные требования к структуре
            if ws.max_row < 2:
                return False, f"{file}: Слишком мало строк"
            if ws.max_column < 34:  # 3 служебных колонки + 31 день
                return False, f"{file}: Слишком мало колонок"

            # Проверяем наличие данных в служебных колонках
            if not ws.cell(row=1, column=3).value:
                return False, f"{file}: Отсутствует заголовок колонки логина"

            wb.close()
        return True, "OK"
    except Exception as e:
        return False, str(e)


async def test_date_handling() -> Tuple[bool, str]:
    """Тестирование обработки дат"""
    try:
        current_date = datetime.now()
        test_dates = [
            current_date + timedelta(days=1),
            current_date + timedelta(days=2),
            current_date + timedelta(days=3),
            current_date.replace(day=1) + timedelta(days=32)
        ]

        for date in test_dates:
            if date.weekday() not in range(7):
                return False, f"Неверный день недели для {date}"
            if date.strftime('%A').lower() not in WEEKDAYS:
                return False, f"День недели не найден в словаре для {date}"

        return True, "OK"
    except Exception as e:
        return False, str(e)


@router.message(Command("test_all"))
async def cmd_test_all(message: Message):
    """Запуск всех тестов"""
    if str(message.from_user.id) not in ADMIN_USERS:
        return

    results = []

    # Выполняем все тесты
    tests = [
        ("📁 Проверка файлов", test_files_existence()),
        ("📊 Проверка структуры", test_schedule_structure()),
        ("📅 Проверка обработки дат", test_date_handling()),
        ("🔄 Проверка команды shift", test_shift_command())
    ]

    for test_name, test_coro in tests:
        success, message_text = await test_coro
        status = "✅" if success else "❌"
        result = f"{status} {test_name}"
        if not success:
            result += f"\n    Ошибка: {message_text}"
        results.append(result)

    # Формируем итоговый отчет
    report = "🧪 Результаты тестирования:\n\n" + "\n\n".join(results)

    # Добавляем общий статус
    all_passed = all(r.startswith("✅") for r in results)
    report += f"\n\n{'✅ Все тесты пройдены' if all_passed else '❌ Есть ошибки в тестах'}"

    await message.answer(report)