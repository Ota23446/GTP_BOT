import os
import logging
import openpyxl
from datetime import datetime, timedelta
from aiogram.utils.keyboard import InlineKeyboardBuilder
from keyboards import show_schedule_choice
from config import ADMIN_USERS
from utils import (
    UserDataManager,
    NOTIFICATION_TYPES,
    format_notification_status,
    is_valid_username
)
from keyboards import show_dates_for_month
from services import get_next_shift, calculate_worked_time
from aiogram import Router, F
from aiogram.filters import Command, CommandStart
from aiogram.types import Message, CallbackQuery
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from keyboards import get_main_keyboard, get_settings_keyboard, get_notification_settings_keyboard
from keyboards import show_specific_date_buttons
from services import get_shift_for_date
from aiogram import Bot


# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

router = Router()
user_manager = UserDataManager()

class NotificationStates(StatesGroup):
    waiting_for_time = State()
    waiting_for_username = State()


@router.message(CommandStart())
async def cmd_start(message: Message):
    username = user_manager.get_user_by_telegram_id(str(message.from_user.id))
    if username:
        await message.answer(
            f"👋 Добро пожаловать!\n"
            f"Ваш логин: {username}",
            reply_markup=get_main_keyboard()  # Используем инлайн-клавиатуру
        )
    else:
        await message.answer(
            "👋 Добро пожаловать!\n"
            "Для начала работы необходимо привязать ваш логин.\n"
            "Используйте команду /register sm_username"
        )


@router.callback_query(F.data == "settings")
async def show_settings(callback: CallbackQuery):
    """Обработка нажатия кнопки Настройки"""
    try:
        username = user_manager.get_user_by_telegram_id(str(callback.from_user.id))
        if not username:
            await callback.message.answer("Сначала необходимо привязать username через /register")
            await callback.answer()
            return

        await callback.message.edit_text(
            "⚙️ Настройки:",
            reply_markup=get_settings_keyboard()
        )
        await callback.answer()

    except Exception as e:
        logging.error(f"Error in show_settings: {e}")
        await callback.answer("Произошла ошибка при открытии настроек")


@router.callback_query(F.data == "shifts")
async def show_shifts(callback: CallbackQuery):
    """Обработка нажатия кнопки Ближайшие смены"""
    try:
        username = user_manager.get_user_by_telegram_id(str(callback.from_user.id))
        if not username:
            await callback.message.edit_text(
                "⚠️ Для просмотра смен необходимо привязать логин.\n"
                "Используйте команду /register sm_username",
                reply_markup=get_main_keyboard()
            )
            await callback.answer()
            return

        shifts = await get_next_shift(username)
        # Отправляем сообщение со сменами и кнопками меню
        await callback.message.edit_text(
            shifts,
            reply_markup=get_main_keyboard()
        )
        await callback.answer()

    except Exception as e:
        logging.error(f"Error in show_shifts: {e}")
        await callback.message.edit_text(
            "Произошла ошибка при получении информации о сменах",
            reply_markup=get_main_keyboard()
        )
        await callback.answer()


@router.callback_query(F.data == "worked_time")
async def show_worked_time(callback: CallbackQuery):
    """Обработка нажатия кнопки Учёт рабочего времени"""
    try:
        username = user_manager.get_user_by_telegram_id(str(callback.from_user.id))
        if not username:
            await callback.message.edit_text(
                "⚠️ Для просмотра рабочего времени необходимо привязать логин.\n"
                "Используйте команду /register sm_username",
                reply_markup=get_main_keyboard()
            )
            await callback.answer()
            return

        worked_time = await calculate_worked_time(username)
        await callback.message.edit_text(
            worked_time,
            reply_markup=get_main_keyboard()
        )
        await callback.answer()

    except Exception as e:
        logging.error(f"Error in show_worked_time: {e}")
        await callback.message.edit_text(
            "Произошла ошибка при подсчете рабочего времени",
            reply_markup=get_main_keyboard()
        )
        await callback.answer()


@router.callback_query(F.data == "help")
async def show_help(callback: CallbackQuery):
    """Обработка нажатия кнопки Помощь"""
    try:
        help_text = (
            "🤖 <b>Возможности бота:</b>\n\n"
            "🗓 <b>Ближайшие смены</b>\n"
            "• Просмотр графика на ближайшие дни\n"
            "• Время начала и окончания смен\n\n"
            "⏱ <b>Учёт рабочего времени</b>\n"
            "• Подсчёт отработанных часов\n"
            "• Статистика по сменам\n\n"
            "⚙️ <b>Настройки</b>\n"
            "• Настройка уведомлений\n"
            "• Выбор времени оповещений\n"
            "• Управление учётной записью\n\n"
            "📝 <b>Основные команды:</b>\n"
            "/register - привязка логина\n"
            "/start - перезапуск бота"
        )
        await callback.message.edit_text(
            help_text,
            parse_mode="HTML",
            reply_markup=get_main_keyboard()  # Добавляем кнопки меню
        )
        await callback.answer()

    except Exception as e:
        logging.error(f"Error in show_help: {e}")
        await callback.answer("Произошла ошибка при открытии помощи")


# Обработчики callback-запросов для inline-кнопок
@router.callback_query(F.data == "set_time")
async def process_time_setting(callback: CallbackQuery, state: FSMContext):
    await state.set_state(NotificationStates.waiting_for_time)
    await callback.message.answer("⏰ Введите желаемое время уведомлений в формате ЧЧ:ММ (например, 19:00)")
    await callback.answer()


@router.callback_query(F.data == "notifications")
async def process_notifications_setting(callback: CallbackQuery):
    await callback.message.edit_text(
        "🔔 Настройка уведомлений:",
        reply_markup=get_notification_settings_keyboard()
    )


@router.callback_query(F.data == "status")
async def process_status(callback: CallbackQuery):
    username = user_manager.get_user_by_telegram_id(str(callback.from_user.id))
    if not username:
        await callback.answer("Сначала привяжите username!")
        return

    user_data = user_manager.get_user_settings(username)
    status_text = (
        f"👤 Username: {username}\n"
        f"⏰ Время уведомлений: {user_data['notification_time']}\n\n"
        "🔔 Статус уведомлений:\n"
    )
    for notif_key, notif_name in NOTIFICATION_TYPES.items():
        status = user_data["notifications"][notif_key]
        status_text += f"{'✅' if status else '❌'} {notif_name}\n"

    await callback.message.edit_text(
        status_text,
        reply_markup=get_settings_keyboard()
    )


@router.callback_query(F.data == "main_menu")
async def process_return_to_main(callback: CallbackQuery):
    await callback.message.delete()
    await callback.message.answer(
        "Главное меню",
        reply_markup=get_main_keyboard()
    )


@router.message(Command("help"))
async def cmd_help(message: Message):
    help_text = (
        "Доступные команды:\n"
        "/start - Начать работу с ботом\n"
        "/register sm_username - Привязать username\n"
        "/settings - Настройки уведомлений\n"
        "/time - Изменить время уведомлений\n"
        "/status - Просмотр текущих настроек\n"
        "/shift - Информация о следующей смене\n"
        "/worked_time - Подсчет отработанного времени"
    )
    await message.answer(help_text)

# Регистрация пользователя
class RegistrationStates(StatesGroup):
    waiting_for_username = State()

@router.message(Command("register"))
async def cmd_register(message: Message, state: FSMContext):
    """Начало регистрации пользователя"""
    user_manager = UserDataManager()
    existing_username = user_manager.get_user_by_telegram_id(str(message.from_user.id))
    if existing_username:
        await message.answer(f"Вы уже зарегистрированы с логином: {existing_username}")
        return

    await state.set_state(RegistrationStates.waiting_for_username)
    await message.answer("Пожалуйста, введите ваш логин в формате sm_username:")

@router.message(RegistrationStates.waiting_for_username)
async def process_username(message: Message, state: FSMContext, bot: Bot):
    """Обработка введенного логина"""
    try:
        username = message.text.lower()
        if not await is_valid_username(username):
            await message.answer("Неверный формат username. Должен начинаться с 'sm_' и быть в нижнем регистре.")
            return

        user_manager = UserDataManager()
        data = user_manager.load_user_data()

        if username in data:
            await message.answer("Этот логин уже зарегистрирован. Попробуйте другой.")
            return

        user_data = {
            "user_id": str(message.from_user.id),
            "notifications": {
                "shift1": True,
                "shift2": True,
                "shift3": True,
                "weekend": True,
                "dayoff": True
            },
            "notification_time": "18:00"
        }

        data[username] = user_data
        user_manager.save_user_data(data)

        await message.answer(
            "✅ Username успешно привязан!\n"
            f"Логин: {username}",
            reply_markup=get_main_keyboard()
        )
        await state.clear()

    except Exception as e:
        logging.error(f"Error in registration: {e}")
        await message.answer("Произошла ошибка при регистрации. Попробуйте позже.")
        await state.clear()


# Настройки уведомлений
@router.message(Command("settings"))
async def cmd_settings(message: Message):
    username = user_manager.get_user_by_telegram_id(str(message.from_user.id))
    if not username:
        await message.answer("Сначала необходимо привязать username через /register")
        return

    builder = InlineKeyboardBuilder()
    user_data = user_manager.get_user_settings(username)

    for notif_key, notif_name in NOTIFICATION_TYPES.items():
        status = user_data["notifications"][notif_key]
        builder.button(
            text=f"{notif_name}: {format_notification_status(status)}",
            callback_data=f"toggle_{notif_key}"
        )
    builder.adjust(1)

    await message.answer(
        "Настройки уведомлений:",
        reply_markup=builder.as_markup()
    )


# Обработка нажатий на кнопки настроек
@router.callback_query(F.data.startswith("toggle_"))
async def process_notification_toggle(callback: CallbackQuery):
    """Обработка переключения статуса уведомлений"""
    try:
        # Проверка наличия username
        username = user_manager.get_user_by_telegram_id(str(callback.from_user.id))
        if not username:
            await callback.answer("Сначала привяжите username!", show_alert=True)
            return

        # Получение типа уведомления из callback_data
        notif_type = callback.data.split("_")[1]

        # Проверяем, что тип уведомления существует в константах
        if notif_type not in NOTIFICATION_TYPES:
            await callback.answer(f"Неизвестный тип уведомления: {notif_type}", show_alert=True)
            return

        user_data = user_manager.get_user_settings(username)

        # Проверяем существование ключа в настройках пользователя
        if notif_type not in user_data["notifications"]:
            # Если ключа нет, инициализируем его
            user_data["notifications"][notif_type] = True
            user_manager.save_user_data(user_data)

        # Инвертируем статус уведомления
        current_status = user_data["notifications"][notif_type]
        user_manager.update_user_notifications(username, notif_type, not current_status)

        # Получаем обновленные данные
        user_data = user_manager.get_user_settings(username)

        # Формируем текст статуса
        notification_text = "🔔 Настройка уведомлений:\n\n"
        for key, name in NOTIFICATION_TYPES.items():
            status = user_data["notifications"].get(key, False)
            notification_text += f"{'✅' if status else '❌'} {name}\n"

        # Обновляем сообщение
        await callback.message.edit_text(
            notification_text,
            reply_markup=get_notification_settings_keyboard()
        )

        # Показываем уведомление о смене статуса
        status_text = "включены" if not current_status else "отключены"
        await callback.answer(
            f"Уведомления '{NOTIFICATION_TYPES[notif_type]}' {status_text}",
            show_alert=False
        )

    except Exception as e:
        logging.error(f"Error in process_notification_toggle: {e}")
        await callback.answer("Произошла ошибка при обновлении настроек", show_alert=True)

@router.callback_query(F.data == "back_to_settings")
async def process_back_to_settings(callback: CallbackQuery):
    """Возврат к основным настройкам"""
    await callback.message.edit_text(
        "⚙️ Настройки:",
        reply_markup=get_settings_keyboard()
    )


# Изменение времени уведомлений
@router.message(Command("time"))
async def cmd_time(message: Message, state: FSMContext):
    username = user_manager.get_user_by_telegram_id(str(message.from_user.id))
    if not username:
        await message.answer("Сначала необходимо привязать username через /register")
        return

    await state.set_state(NotificationStates.waiting_for_time)
    await message.answer(
        "Введите желаемое время уведомлений в формате ЧЧ:ММ (например, 19:00)"
    )


@router.message(NotificationStates.waiting_for_time)
async def process_notification_time(message: Message, state: FSMContext):
    """Обработка установки времени уведомлений"""
    username = user_manager.get_user_by_telegram_id(str(message.from_user.id))

    if user_manager.update_notification_time(username, message.text):
        await message.answer(
            f"⏰ Время уведомлений успешно обновлено на {message.text}",
            reply_markup=get_settings_keyboard()
        )
    else:
        await message.answer(
            "❌ Неверный формат времени. Используйте формат ЧЧ:ММ (например, 19:00)",
            reply_markup=get_settings_keyboard()
        )

    await state.clear()

@router.callback_query(F.data == "change_username")
async def process_change_username(callback: CallbackQuery, state: FSMContext):
    """Начало процесса изменения username"""
    await state.set_state(NotificationStates.waiting_for_username)
    await callback.message.answer(
        "👤 Введите новый username в формате sm_username\n"
        "Например: sm_ivan"
    )
    await callback.answer()


class NotificationStates(StatesGroup):
    waiting_for_time = State()
    waiting_for_username = State()


@router.message(NotificationStates.waiting_for_username)
async def process_new_username(message: Message, state: FSMContext):
    """Обработка нового username"""
    new_username = message.text.lower()

    if not await is_valid_username(new_username):
        await message.answer(
            "❌ Неверный формат username. Должен начинаться с 'sm_' и быть в нижнем регистре.",
            reply_markup=get_settings_keyboard()
        )
        await state.clear()
        return

    data = user_manager.load_user_data()
    if new_username not in data:
        await message.answer(
            "❌ Такой username не найден в системе.",
            reply_markup=get_settings_keyboard()
        )
        await state.clear()
        return

    # Проверяем, не привязан ли уже этот username к другому пользователю
    if data[new_username]["user_id"] != "0" and data[new_username]["user_id"] != str(message.from_user.id):
        await message.answer(
            "❌ Этот username уже привязан к другому пользователю.",
            reply_markup=get_settings_keyboard()
        )
        await state.clear()
        return

    # Обновляем username
    data[new_username]["user_id"] = str(message.from_user.id)
    user_manager.save_user_data(data)

    await message.answer(
        f"✅ Username успешно изменен на {new_username}",
        reply_markup=get_settings_keyboard()
    )
    await state.clear()

@router.message(Command("status"))
async def cmd_status(message: Message):
    username = user_manager.get_user_by_telegram_id(str(message.from_user.id))
    if not username:
        await message.answer("Сначала необходимо привязать username через /register")
        return

    user_data = user_manager.get_user_settings(username)
    if not user_data:
        await message.answer("Ошибка: данные пользователя не найдены")
        return

    status_text = [
        f"📱 Username: {username}",
        f"⏰ Время уведомлений: {user_data['notification_time']}",
        "\n📋 Статус уведомлений:"
    ]

    for notif_key, notif_name in NOTIFICATION_TYPES.items():
        status = user_data["notifications"][notif_key]
        status_text.append(f"- {notif_name}: {format_notification_status(status)}")

    await message.answer("\n".join(status_text))

@router.message(Command("shift"))
async def cmd_shift(message: Message):
    username = user_manager.get_user_by_telegram_id(str(message.from_user.id))
    if not username:
        await message.answer("Сначала необходимо привязать username через /register")
        return

    try:
        next_shift_info = await get_next_shift(username)
        await message.answer(next_shift_info)
    except Exception as e:
        logging.error(f"Error in shift command: {e}")
        await message.answer("Произошла ошибка при получении информации о смене")

@router.message(Command("worked_time"))
async def cmd_worked_time(message: Message):
    username = user_manager.get_user_by_telegram_id(str(message.from_user.id))
    if not username:
        await message.answer("Сначала необходимо привязать username через /register")
        return

    try:
        worked_time = await calculate_worked_time(username)
        await message.answer(worked_time)
    except Exception as e:
        logging.error(f"Error in worked_time command: {e}")
        await message.answer("Произошла ошибка при подсчете отработанного времени")


@router.message(Command("debug_schedule"))
async def cmd_debug_schedule(message: Message):
    if str(message.from_user.id) not in ADMIN_USERS:
        return

    try:
        wb = openpyxl.load_workbook('schedule.xlsx')
        ws = wb.active

        debug_info = ["📋 Отладочная информация о расписании:"]

        # Получаем значение ячейки для проверки
        username = "sm_kirillts"
        current_date = datetime.now()
        next_date = current_date + timedelta(days=1)
        next_day_col = next_date.day + 1

        debug_info.append(f"\n📅 Текущая дата: {current_date.strftime('%d.%m.%Y')}")
        debug_info.append(f"📅 Следующая дата: {next_date.strftime('%d.%m.%Y')}")
        debug_info.append(f"📊 Номер колонки для следующей даты: {next_day_col}")

        found = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).lower() == username.lower():
                    shift = ws.cell(row=cell.row, column=next_day_col).value
                    debug_info.append(f"\n👤 Найден пользователь: {username}")
                    debug_info.append(f"📍 Строка в таблице: {cell.row}")
                    debug_info.append(f"🔄 Значение смены: {shift}")
                    found = True
                    break
            if found:
                break

        if not found:
            debug_info.append(f"\n❌ Пользователь {username} не найден в расписании")

        # Добавим информацию о файле
        debug_info.append(f"\n📁 Путь к файлу: {os.path.abspath('schedule.xlsx')}")
        debug_info.append(f"📊 Размер таблицы: {ws.max_row}x{ws.max_column}")

        await message.answer("\n".join(debug_info))
        wb.close()
    except Exception as e:
        error_msg = f"🚫 Ошибка при отладке:\n{str(e)}"
        if 'schedule.xlsx' not in os.listdir():
            error_msg += "\n\nФайл schedule.xlsx не найден в директории!"
        await message.answer(error_msg)


@router.message(Command("check_cell"))
async def cmd_check_cell(message: Message):
    if str(message.from_user.id) not in ADMIN_USERS:
        return

    try:
        wb = openpyxl.load_workbook('schedule.xlsx')
        ws = wb.active

        # Проверяем конкретную ячейку
        next_date = datetime.now() + timedelta(days=1)
        next_day_col = next_date.day + 3  # +3 так как смены начинаются с 4-й колонки

        debug_info = [
            "🔍 Проверка ячеек:",
            f"📅 Завтрашняя дата: {next_date.strftime('%d.%m.%Y')}",
            f"📊 Номер колонки: {next_day_col}\n"
        ]

        # Проверяем все ячейки в колонке с логинами (3-я колонка)
        for row in range(1, ws.max_row + 1):
            login_cell = ws.cell(row=row, column=3)  # 3-я колонка для логинов
            shift_cell = ws.cell(row=row, column=next_day_col)

            if login_cell.value:
                debug_info.append(f"Строка {row}: {login_cell.value} -> {shift_cell.value}")
                if str(login_cell.value).lower() == 'sm_kirillts':
                    debug_info.append(f"\n🎯 Найдена целевая строка!")
                    debug_info.append(f"Координаты ячейки смены: {shift_cell.coordinate}")
                    debug_info.append(f"Тип данных: {type(shift_cell.value)}")
                    debug_info.append(f"Значение: '{shift_cell.value}'")

        await message.answer("\n".join(debug_info))
        wb.close()
    except Exception as e:
        await message.answer(f"🚫 Ошибка при проверке: {str(e)}")

async def process_schedule_day(callback: CallbackQuery, month: int, day: int):
    """Обрабатывает выбор конкретной даты и показывает смену"""
    user_id = str(callback.from_user.id)
    username = user_manager.get_user_by_telegram_id(user_id)
    if not username:
        await callback.message.answer("Сначала зарегистрируйтесь! /register")
        return

    current_date = datetime.now()
    year = current_date.year
    if month > 12:
        year += 1

    target_date = datetime(year, month, day)

    filename = "schedule.xlsx" if month == current_date.month else "schedule_next.xlsx"
    shift_info = await get_shift_for_date(username, target_date, filename)
    await callback.message.answer(shift_info)
    await callback.message.answer("Главное меню:", reply_markup=get_main_keyboard())

@router.callback_query(F.data.startswith("schedule_month:"))
async def callback_schedule_month(callback: CallbackQuery):
    month = int(callback.data.split(":")[1])
    await show_dates_for_month(callback, month)


@router.callback_query(F.data.startswith("schedule_dates:"))
async def callback_schedule_dates(callback: CallbackQuery):
    _, month_str, days_range = callback.data.split(":")
    month = int(month_str)
    await show_specific_date_buttons(callback, month, days_range)


@router.callback_query(F.data.startswith("schedule_day:"))
async def callback_schedule_day(callback: CallbackQuery):
    _, month_str, day_str = callback.data.split(":")
    month = int(month_str)
    day = int(day_str)
    await process_schedule_day(callback, month, day)

@router.callback_query(F.data == "check_shift")
async def callback_check_shift(callback: CallbackQuery):
    await show_schedule_choice(callback)