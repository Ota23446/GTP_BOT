import telebot
import json
import logging
import tenacity
import openpyxl
from datetime import datetime, timedelta
from threading import Thread
import schedule
import time
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton
import requests
from docx import Document
import openpyxl
import os
from datetime import datetime
import win32com.client
import calendar
import locale
import pythoncom
import psutil  # Добавляем импорт psutil
import sys
import atexit
import logging
import re  # Импортируем модуль re
from datetime import datetime, timedelta
import openpyxl
import holidays
import win32com.client
import os
from pathlib import Path
import shutil
from dateutil.relativedelta import relativedelta
import locale
from dotenv import load_dotenv
locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')


# Настройка логирования
logging.basicConfig(level=logging.INFO, filename='bot.log',
                    format='%(asctime)s - %(levelname)s - %(message)s')
# Загружаем переменные из .env
load_dotenv()
BOT_TOKEN = os.getenv('BOT_TOKEN')
ADMIN_USERS = set(os.getenv('ADMIN_USERS').split(','))
EXCEL_FILE = 'schedule.xlsx'
JSON_FILE = 'user_data.json'
MONTHS_RU = {
    1: "Январь",
    2: "Февраль",
    3: "Март",
    4: "Апрель",
    5: "Май",
    6: "Июнь",
    7: "Июль",
    8: "Август",
    9: "Сентябрь",
    10: "Октябрь",
    11: "Ноябрь",
    12: "Декабрь"
}


bot = telebot.TeleBot(BOT_TOKEN)

def get_current_month_year():
    # Установка русской локали
    locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')
    current_date = datetime.now()
    # Получаем название месяца с большой буквы
    month = current_date.strftime('%B').capitalize()
    year = current_date.strftime('%Y')
    return f"{month} {year}"

def download_and_process_doc():
    url = "http://confluence.jira.lan:8090/exportword?pageId=24577712"
    doc_folder = "documents"
    excel_folder = "excel"
    main_schedule = 'schedule.xlsx'  # Файл для текущего месяца
    next_schedule = 'schedule_next.xlsx'  # Файл для следующего месяца

    # Создаем папки
    for folder in [doc_folder, excel_folder]:
        if not os.path.exists(folder):
            os.makedirs(folder)

    try:
        # Инициализация COM
        pythoncom.CoInitialize()

        # Скачивание файла
        response = requests.get(url, verify=False)
        response.raise_for_status()

        # Сохранение .doc файла
        current_date = datetime.now()
        # current_date = datetime(2024, 11, 1)  # Для тестирования
        current_date_str = current_date.strftime("%Y%m%d")
        doc_filename = f'schedule_{current_date_str}.doc'
        doc_path = os.path.join(doc_folder, doc_filename)
        doc_path_abs = os.path.abspath(doc_path)

        with open(doc_path, 'wb') as f:
            f.write(response.content)

        # Открываем Word
        word = win32com.client.Dispatch('Word.Application')
        word.Visible = False

        try:
            # Открываем документ
            doc = word.Documents.Open(doc_path_abs)

            # Получаем текущий и следующий месяц
            next_date = current_date + relativedelta(months=1)
            # next_date = datetime(2024, 12, 1)  # Для тестирования
            target_month = MONTHS_RU[current_date.month]
            target_year = current_date.strftime('%Y')
            target_header = f"{target_month} {target_year}"

            next_month = MONTHS_RU[next_date.month]
            next_year = next_date.strftime('%Y')
            next_header = f"{next_month} {next_year}"

            logging.info(f"Looking for table with header: {target_header}")
            logging.info(f"Looking for table with header: {next_header}")

            # Создаем Excel файлы
            wb_current = openpyxl.Workbook()
            ws_current = wb_current.active
            wb_next = openpyxl.Workbook()
            ws_next = wb_next.active

            # Ищем нужные таблицы
            found_current_table = False
            found_next_table = False
            content = doc.Content
            tables = doc.Tables

            for i in range(1, tables.Count + 1):
                table = tables.Item(i)

                # Получаем текст перед таблицей
                if i == 1:
                    range_start = 0
                else:
                    range_start = tables.Item(i - 1).Range.End
                range_end = table.Range.Start

                header_text = doc.Range(range_start, range_end).Text

                if target_header in header_text:
                    found_current_table = True
                    # Копируем данные из таблицы
                    rows = table.Rows.Count
                    cols = table.Columns.Count

                    for row in range(1, rows + 1):
                        for col in range(1, cols + 1):
                            try:
                                # Получаем значение ячейки
                                cell_value = table.Cell(Row=row, Column=col).Range.Text
                                # Очищаем текст
                                cell_value = cell_value.strip().replace('\r', '').replace('\a', '')
                                # Записываем в Excel
                                ws_current.cell(row=row, column=col, value=cell_value)
                            except Exception as e:
                                logging.warning(f"Error processing cell [{row},{col}]: {e}")

                elif next_header in header_text:
                    found_next_table = True
                    # Копируем данные из таблицы
                    rows = table.Rows.Count
                    cols = table.Columns.Count

                    for row in range(1, rows + 1):
                        for col in range(1, cols + 1):
                            try:
                                # Получаем значение ячейки
                                cell_value = table.Cell(Row=row, Column=col).Range.Text
                                # Очищаем текст
                                cell_value = cell_value.strip().replace('\r', '').replace('\a', '')
                                # Записываем в Excel
                                ws_next.cell(row=row, column=col, value=cell_value)
                            except Exception as e:
                                logging.warning(f"Error processing cell [{row},{col}]: {e}")

            try:
                # Сохраняем файл текущего месяца
                excel_filename = f'schedule_{current_date_str}.xlsx'
                excel_path = os.path.join(excel_folder, excel_filename)
                wb_current.save(excel_path)
                wb_current.close()  # Добавили close workbook
            except Exception as e:
                logging.error(f"Error saving current schedule file: {e}")

            try:
                # Сохраняем файл следующего месяца
                next_excel_path = os.path.join(excel_folder, next_schedule)
                wb_next.save(next_excel_path)
                wb_next.close()  # Добавили close workbook
            except Exception as e:
                logging.error(f"Error saving next schedule file: {e}")

            # Обновляем общий файл в корне проекта
            try:
                if os.path.exists(main_schedule):
                    os.remove(main_schedule)
                wb_current.save(main_schedule)
                logging.info(f"Successfully updated main schedule: {main_schedule}")
            except Exception as e:
                logging.error(f"Error updating main schedule file: {e}")

            # Очистка папок после успешного обновления
            if found_current_table or found_next_table:
                for folder in [doc_folder, excel_folder]:
                    for filename in os.listdir(folder):
                        file_path = os.path.join(folder, filename)
                        try:
                            if os.path.isfile(file_path):
                                os.unlink(file_path)  # Удаляем файлы
                                logging.info(f"Deleted file: {file_path}")
                        except Exception as e:
                            logging.warning(f"Error deleting file {file_path}: {e}")

            # Возвращаем результат
            current_status = "Текущий месяц обновлен" if found_current_table else "Текущий месяц не найден"
            next_status = "Следующий месяц обновлен" if found_next_table else "Следующий месяц не найден"
            return f"{current_status}. {next_status}."

        except Exception as e:
            logging.error(f"Error processing document: {e}")
            logging.exception("Full traceback:")
            return "Произошла ошибка при обработке документа."
        finally:
            try:
                doc.Close(SaveChanges=False)
            except:
                pass
            try:
                word.Quit()
            except:
                pass

    except Exception as e:
        logging.error(f"Error in main process: {e}")
        logging.exception("Full traceback:")
        return "Произошла ошибка в основном процессе."
    finally:
        try:
            word.Quit()
        except:
            pass
        pythoncom.CoUninitialize()

@bot.message_handler(commands=['debug'])
def debug_command(message):
    user_id = str(message.from_user.id)
    if user_id in ADMIN_USERS:
        try:
            current_date = datetime.now()
            target_month = MONTHS_RU[current_date.month]
            target_year = current_date.strftime('%Y')
            target_header = f"{target_month} {target_year}"

            doc_path = os.path.join("documents", f'schedule_{current_date.strftime("%Y%m%d")}.doc')

            bot.reply_to(message,
                         f"Текущая дата: {current_date.strftime('%d.%m.%Y')}\n"
                         f"Искомый заголовок: {target_header}\n"
                         f"Путь к документу: {doc_path}\n"
                         f"Файл существует: {os.path.exists(doc_path)}")

            if os.path.exists(doc_path):
                file_size = os.path.getsize(doc_path)
                bot.reply_to(message, f"Размер файла: {file_size} байт")

        except Exception as e:
            bot.reply_to(message, f"Ошибка отладки: {str(e)}")



def fix_permissions():
    """Fix permissions for all necessary folders and files."""
    folders = ['documents', 'excel']
    try:
        for folder in folders:
            if not os.path.exists(folder):
                os.makedirs(folder, mode=0o777)
            else:
                os.chmod(folder, 0o777)

            # Fix permissions for all files in the folder
            for filename in os.listdir(folder):
                file_path = os.path.join(folder, filename)
                try:
                    os.chmod(file_path, 0o666)
                except Exception as e:
                    logging.error(f"Error fixing permissions for {file_path}: {e}")

        return True
    except Exception as e:
        logging.error(f"Error fixing permissions: {e}")
        return False

def update_schedule():
    new_excel_file = download_and_process_doc()
    if new_excel_file:
        global EXCEL_FILE
        EXCEL_FILE = new_excel_file
        # Обновление данных
        excel_logins = load_excel_logins()
        schedule_common_jobs()
        schedule_updates()
        logging.info("Schedule successfully updated")
    else:
        logging.error("Failed to update schedule")

def load_data(filepath, default_value, load_function):
    try:
        with open(filepath, 'r') as f:
            content = f.read()
            if content:
                return load_function(content)
            else:
                return default_value
    except FileNotFoundError:
        return default_value
    except Exception as e:
        logging.error(f"Error loading data from {filepath}: {e}")
        return default_value



def load_user_data():
    data = load_data(JSON_FILE, {}, json.loads)
    for login, user_data_entry in data.items():
        if not isinstance(user_data_entry, dict):
            data[login] = {
                "user_id": user_data_entry,
                "notifications": {
                    "shift1": True,
                    "shift2": True,
                    "shift3": True,
                    "weekend_duty": True,
                    "day_off": True
                },
                "notification_time": "18:00"
            }
    return data


def send_shift_notifications():
    today = datetime.now().date()
    tomorrow = today + timedelta(days=1)
    current_time = datetime.now().strftime("%H:%M")

    # Создаем множество обработанных пользователей
    processed_users = set()

    for login, data in user_data.items():
        user_id = data.get('user_id')
        if user_id in processed_users:
            continue  # Пропускаем уже обработанных пользователей

        notifications = data.get('notifications', {})
        notification_time = data.get('notification_time', "18:00")

        if current_time == notification_time:
            shifts, _ = get_user_info(login)
            if shifts:
                tomorrow_shift = next((shift for date, shift in shifts if date == tomorrow), None)

                if tomorrow_shift == 1 and notifications.get('shift1'):
                    send_notification(user_id, f"Напоминание: завтра {tomorrow.strftime('%d.%m.%Y')} у вас 1 смена.")
                elif tomorrow_shift == 2 and notifications.get('shift2'):
                    send_notification(user_id, f"Напоминание: завтра {tomorrow.strftime('%d.%m.%Y')} у вас 2 смена.")
                elif tomorrow_shift == 3 and notifications.get('shift3'):
                    send_notification(user_id, f"Напоминание: завтра {tomorrow.strftime('%d.%m.%Y')} у вас 3 смена.")
                elif str(tomorrow_shift).upper() in ['В', 'B'] and notifications.get('weekend_duty'):
                    send_notification(user_id, f"Напоминание: завтра {tomorrow.strftime('%d.%m.%Y')} вы дежурите.")
                elif (tomorrow_shift is None or tomorrow_shift == '') and notifications.get('day_off'):
                    send_notification(user_id, f"Напоминание: завтра {tomorrow.strftime('%d.%m.%Y')} у вас выходной.")

            processed_users.add(user_id)  # Добавляем пользователя в множество обработанных

def send_monday_notification():
    processed_users = set()
    for login, data in user_data.items():
        user_id = data.get('user_id')
        if user_id and user_id not in processed_users:
            send_notification(user_id, "Завтра вторник, нужно проверить битрикс на просроченные задачи.")
            processed_users.add(user_id)

def send_tuesday_notification():
    processed_users = set()
    for login, data in user_data.items():
        user_id = data.get('user_id')
        if user_id and user_id not in processed_users:
            send_notification(user_id, "Сегодня вторник, нужно проверить битрикс на просроченные задачи.")
            processed_users.add(user_id)

def send_notification(user_id, message):
    try:
        bot.send_message(user_id, message)
    except Exception as e:
        logging.error(f"Failed to send notification to user {user_id}: {e}")


# Функция для настройки расписания уведомлений для пользователя
def schedule_jobs_for_user(user_id, notification_time):
    # Очищаем предыдущие задачи для этого пользователя
    schedule.clear(str(user_id))


    # Больше не планируем отдельные уведомления для каждого пользователя
    # Вместо этого обновляем время в данных пользователя
    for login, data in user_data.items():
        if data.get('user_id') == user_id:
            data['notification_time'] = notification_time
            save_user_data(user_data)
            break

def send_worked_time_notification():
    for login, data in user_data.items():
        user_id = data.get('user_id')
        worked_time = calculate_worked_time(login)
        bot.send_message(user_id, f"Отработанное время с начала месяца: {worked_time}")


@bot.message_handler(commands=['worked_time'])
def worked_time_command(message):
    try:
        user_id = str(message.from_user.id)
        # Находим логин пользователя
        login = None
        for l, data in user_data.items():  # user_data должен быть определен где-то выше
            if str(data.get('user_id')) == user_id:
                login = l
                break

        if login:
            # Добавляем сообщение о начале подсчета
            processing_msg = bot.reply_to(message, "⌛ Подсчитываю отработанное время...")

            # Определяем имя Excel-файла (адаптируйте под вашу логику)
            excel_file = "schedule.xlsx"  # Или, например, f"{login}.xlsx", если имя файла зависит от login

            # Получаем отчет
            worked_time = calculate_worked_time(login, excel_file)

            # Удаляем сообщение о подсчете
            bot.delete_message(message.chat.id, processing_msg.message_id)

            # Отправляем результат
            bot.reply_to(message, worked_time)
        else:
            bot.reply_to(message, "Вы не зарегистрированы в системе. Используйте /start для регистрации.")

    except Exception as e:
        logging.error(f"Error in worked_time_command: {e}")
        bot.reply_to(message, "Произошла ошибка при выполнении команды")

def schedule_common_jobs():
    # Очищаем все существующие задачи
    schedule.clear()

def clear_logs(log_file='bot.log', retention_days=7):
    """
    Очищает файл логов, если с момента последнего изменения прошло больше retention_days.
    """
    try:
        if os.path.exists(log_file):
            # Получаем время последней модификации файла
            last_modified = datetime.fromtimestamp(os.path.getmtime(log_file))
            now = datetime.now()

            # Проверяем, прошло ли больше заданного количества дней
            if (now - last_modified).days >= retention_days:
                # Очищаем файл
                with open(log_file, 'w'):
                    pass  # Просто открываем файл в режиме записи для очистки
                logging.info(f"Лог-файл {log_file} очищен.")
            else:
                logging.info(f"Лог-файл {log_file} не требует очистки. Последняя модификация: {last_modified}")
        else:
            logging.info(f"Лог-файл {log_file} отсутствует.")
    except Exception as e:
        logging.error(f"Ошибка при очистке логов: {e}")


def schedule_updates():
    try:
        # Очищаем все существующие задачи перед планированием новых
        schedule.clear()
        logging.info("Cleared existing schedules")

        # Планируем ежедневное обновление расписания
        schedule.every().day.at("00:01").do(update_schedule)

        # Планируем очистку логов раз в неделю (воскресенье в 03:00)
        schedule.every().sunday.at("03:00").do(clear_logs)

        # Создаем множество уникальных времен уведомлений
        notification_times = set()
        for login, data in user_data.items():
            notification_time = data.get('notification_time', "18:00")
            notification_times.add(notification_time)

        # Планируем уведомления для каждого уникального времени
        for time in notification_times:
            schedule.every().day.at(time).do(send_shift_notifications_for_time, time)
            logging.info(f"Scheduled notifications for {time}")

        # Планируем обязательные уведомления
        schedule.every().monday.at("09:30").do(send_monday_notification)
        schedule.every().tuesday.at("09:30").do(send_tuesday_notification)
        schedule.every().friday.at("09:30").do(send_worked_time_notification)
        logging.info("Scheduled weekly notifications")

        # Логируем все запланированные задачи
        jobs_info = "Scheduled tasks:\n"
        for job in schedule.jobs:
            jobs_info += f"- {job}\n"
        logging.info(jobs_info)

    except Exception as e:
        logging.error(f"Error in schedule_updates: {e}")
        logging.exception("Full traceback:")


@bot.message_handler(commands=['refresh_schedule'])
def refresh_schedule_command(message):
    user_id = str(message.from_user.id)
    if user_id in ADMIN_USERS:
        try:
            schedule_updates()

            # Проверяем запланированные задачи
            jobs_info = "Запланированные задачи:\n"
            for job in schedule.jobs:
                jobs_info += f"- {job}\n"

            bot.reply_to(message, f"Расписание обновлено!\n\n{jobs_info}")

        except Exception as e:
            bot.reply_to(message, f"Ошибка при обновлении расписания: {str(e)}")
    else:
        bot.reply_to(message, "У вас нет прав для выполнения этой команды")


def send_shift_notifications_for_time(target_time):
    current_time = datetime.now().strftime("%H:%M")
    logging.info(f"=== Starting notification check ===")
    logging.info(f"Current time: {current_time}, Target time: {target_time}")

    if current_time != target_time:
        logging.info(f"Times don't match, skipping")
        return

    today = datetime.now().date()
    tomorrow = today + timedelta(days=1)

    logging.info(f"Checking shifts for {tomorrow.strftime('%d.%m.%Y')}")

    for login, data in user_data.items():
        try:
            user_id = data.get('user_id')
            notification_time = data.get('notification_time', "18:00")
            notifications = data.get('notifications', {})

            logging.info(f"\nProcessing user {login}:")
            logging.info(f"User ID: {user_id}")
            logging.info(f"Notification time: {notification_time}")
            logging.info(f"Notification settings: {notifications}")

            if notification_time != target_time:
                logging.info(f"Skipping - different notification time")
                continue

            # Получаем смены для текущего пользователя
            shifts, _ = get_user_info(login)

            if not shifts:
                logging.warning(f"No shifts found in current schedule.")
                continue

            # Ищем смену для завтрашнего дня
            tomorrow_shift = next((shift for date, shift in shifts if date == tomorrow), None)

            # Если завтрашняя дата относится к следующему месяцу, загружаем данные из schedule_next.xlsx
            if tomorrow.month != today.month:
                logging.info("Tomorrow belongs to the next month. Checking schedule_next.xlsx...")
                next_month_shifts, _ = get_user_info_from_next_schedule(login)  # Используем новую функцию
                tomorrow_shift = next((shift for date, shift in next_month_shifts if date == tomorrow), None)

            logging.info(f"Tomorrow's shift: {tomorrow_shift}")

            try:
                # Преобразуем tomorrow_shift в число, если это возможно
                if isinstance(tomorrow_shift, str):
                    if tomorrow_shift.isdigit():
                        tomorrow_shift = int(tomorrow_shift)

                message_sent = False
                if tomorrow_shift == 1 and notifications.get('shift1', True):
                    send_notification(user_id, f"Напоминание: завтра {tomorrow.strftime('%d.%m.%Y')} у вас 1 смена.")
                    message_sent = True
                    logging.info("Sent notification for shift 1")
                elif tomorrow_shift == 2 and notifications.get('shift2', True):
                    send_notification(user_id, f"Напоминание: завтра {tomorrow.strftime('%d.%m.%Y')} у вас 2 смена.")
                    message_sent = True
                    logging.info("Sent notification for shift 2")
                elif tomorrow_shift == 3 and notifications.get('shift3', True):
                    send_notification(user_id, f"Напоминание: завтра {tomorrow.strftime('%d.%m.%Y')} у вас 3 смена.")
                    message_sent = True
                    logging.info("Sent notification for shift 3")
                elif str(tomorrow_shift).upper() in ['В', 'B'] and notifications.get('weekend_duty', True):
                    send_notification(user_id, f"Напоминание: завтра {tomorrow.strftime('%d.%m.%Y')} вы дежурите.")
                    message_sent = True
                    logging.info("Sent notification for weekend duty")
                elif (tomorrow_shift is None or tomorrow_shift == '' or tomorrow_shift == 0) and notifications.get('day_off', True):
                    send_notification(user_id, f"Напоминание: завтра {tomorrow.strftime('%d.%m.%Y')} у вас выходной.")
                    message_sent = True
                    logging.info("Sent notification for day off")

                if not message_sent:
                    logging.info(f"No notification sent. Shift type: {tomorrow_shift}, Type: {type(tomorrow_shift)}")

            except Exception as e:
                logging.error(f"Error sending notification: {e}")
                logging.exception("Full traceback:")

        except Exception as e:
            logging.error(f"Error processing user {login}: {e}")
            logging.exception("Full traceback:")

    logging.info("=== Notification check completed ===")

def get_user_info_from_next_schedule(login):
    """
    Получает информацию о сменах пользователя из расписания следующего месяца.
    """
    schedule_next_data = load_schedule_from_excel("schedule_next.xlsx")  # Загрузка расписания следующего месяца

    # Ищем пользователя в расписании следующего месяца
    next_entry = next((entry for entry in schedule_next_data if entry["login"] == login), None)

    if not next_entry:
        logging.warning(f"Не найдена информация о сменах для логина {login} в следующем месяце.")
        return [], None

    # Генерируем список смен
    next_month = (datetime.now().month % 12) + 1
    next_year = datetime.now().year if next_month > 1 else datetime.now().year + 1
    shifts = []

    for day, shift in enumerate(next_entry["shifts"], start=1):
        try:
            date = datetime(next_year, next_month, day).date()
            shifts.append((date, shift))
        except ValueError:
            logging.warning(f"Некорректная дата: {next_year}-{next_month}-{day}")

    return shifts, next_entry.get("name")

@bot.message_handler(commands=['debug_user'])
def debug_user_command(message):
    user_id = str(message.from_user.id)

    # Выводим информацию о текущем пользователе
    debug_info = f"Ваш ID: {user_id}\n\n"
    debug_info += "Зарегистрированные пользователи:\n"

    # Проверяем все записи в user_data
    for login, data in user_data.items():
        stored_id = str(data.get('user_id', 'None'))
        debug_info += f"Логин: {login}, ID: {stored_id}\n"

    bot.reply_to(message, debug_info)

@bot.message_handler(commands=['check_schedule'])
def check_schedule_command(message):
    user_id = str(message.from_user.id)
    if user_id in ADMIN_USERS:
        try:
            response = "Запланированные задачи:\n\n"
            for job in schedule.jobs:
                response += f"- {job}\n"

            response += "\nТекущие настройки пользователей:\n"
            for login, data in user_data.items():
                notification_time = data.get('notification_time', "18:00")
                response += f"- {login}: {notification_time}\n"

            bot.reply_to(message, response)
        except Exception as e:
            bot.reply_to(message, f"Ошибка при проверке расписания: {str(e)}")

# Функция для запуска планировщика
def run_scheduler():
    logging.info("Scheduler started")
    # Планируем задачи при запуске
    schedule_updates()

    while True:
        try:
            schedule.run_pending()
            time.sleep(1)
        except Exception as e:
            logging.error(f"Error in scheduler: {e}")
            logging.exception("Full traceback:")
            time.sleep(5)

def load_excel_logins():
    """Loads logins from the Excel file."""
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        logins = set()
        for row in sheet.iter_rows(min_row=4, values_only=True):
            login = row[2]  # Колонка C (логин)
            if login:
                logins.add(str(login))
        return logins
    except Exception as e:
        logging.error(f"Error loading logins from Excel: {e}")
        return set()

def save_user_data(data):
    """Saves user data to JSON file."""
    try:
        with open(JSON_FILE, 'w') as f:
            json.dump(data, f, indent=4)
    except IOError as e:
        logging.error(f"Error writing to {JSON_FILE}: {e}")

# Load data on startup
user_data = load_user_data()
excel_logins = load_excel_logins()

# --- Bot Commands ---

@bot.message_handler(commands=['start'])
def start_command(message):
    user_id = str(message.from_user.id)
    if any(user_id == data.get('user_id') for data in user_data.values()):
        bot.reply_to(message, "Вы уже зарегистрированы в системе.")
    else:
        bot.reply_to(message, "Добро пожаловать! Пожалуйста, введите ваш логин из Jira для регистрации.")
        bot.register_next_step_handler(message, process_login)

def process_login(message):
    user_id = str(message.from_user.id)
    login = message.text.strip()
    if login in excel_logins:
        user_data[login] = {
            "user_id": user_id,
            "notifications": {
                "shift1": True,
                "shift2": True,
                "shift3": True,
                "weekend_duty": True,
                "day_off": True
            },
            "notification_time": "18:00"
        }
        save_user_data(user_data)
        schedule_jobs_for_user(user_id, "18:00")
        bot.reply_to(message, "Вы успешно зарегистрированы! По умолчанию уведомления установлены на 18:00")





@bot.message_handler(commands=['settings'])
def settings_command(message):
    user_id = str(message.from_user.id)

    # Добавим отладочный вывод
    logging.info(f"Settings requested by user_id: {user_id}")
    logging.info(f"Current user_data: {json.dumps(user_data, indent=2)}")

    # Ищем логин пользователя
    login = None
    for l, data in user_data.items():
        if str(data.get('user_id')) == user_id:
            login = l
            break

    if login and login in user_data:
        notifications = user_data[login].get("notifications", {
            "shift1": True,
            "shift2": True,
            "shift3": True,
            "weekend_duty": True,
            "day_off": True
        })
        notification_time = user_data[login].get("notification_time", "18:00")

        # Создаем клавиатуру
        markup = InlineKeyboardMarkup()
        markup.add(InlineKeyboardButton(
            f"1 Смена ({'Вкл' if notifications.get('shift1', True) else 'Выкл'})",
            callback_data='toggle_shift1'
        ))
        markup.add(InlineKeyboardButton(
            f"2 Смена ({'Вкл' if notifications.get('shift2', True) else 'Выкл'})",
            callback_data='toggle_shift2'
        ))
        markup.add(InlineKeyboardButton(
            f"3 Смена ({'Вкл' if notifications.get('shift3', True) else 'Выкл'})",
            callback_data='toggle_shift3'
        ))
        markup.add(InlineKeyboardButton(
            f"Дежурство ({'Вкл' if notifications.get('weekend_duty', True) else 'Выкл'})",
            callback_data='toggle_weekend_duty'
        ))
        markup.add(InlineKeyboardButton(
            f"Выходной ({'Вкл' if notifications.get('day_off', True) else 'Выкл'})",
            callback_data='toggle_day_off'
        ))
        markup.add(InlineKeyboardButton(
            f"Время уведомлений: {notification_time}",
            callback_data='set_time'
        ))

        bot.send_message(
            message.chat.id,
            f"Настройки уведомлений для пользователя {login}:",
            reply_markup=markup
        )
    else:
        # Если пользователь не найден, выводим подробную информацию
        error_msg = (
            "Ошибка доступа к настройкам.\n"
            f"Ваш ID: {user_id}\n"
            "Зарегистрированные пользователи:\n"
        )
        for l, data in user_data.items():
            error_msg += f"Логин: {l}, ID: {data.get('user_id')}\n"

        bot.reply_to(message, error_msg)
        logging.error(f"User {user_id} not found in user_data or has invalid data")

@bot.callback_query_handler(func=lambda call: True)
def callback_handler(call):
    user_id = str(call.from_user.id)
    login = next((login for login, data in user_data.items() if data.get('user_id') == user_id), None)

    if login:
        if call.data.startswith('toggle_'):
            setting = call.data.replace('toggle_', '')
            user_data[login]['notifications'][setting] = not user_data[login]['notifications'][setting]
            save_user_data(user_data)

            # Обновляем сообщение с настройками
            notifications = user_data[login]["notifications"]
            notification_time = user_data[login]["notification_time"]
            markup = InlineKeyboardMarkup()
            markup.add(InlineKeyboardButton(
                f"1 Смена ({'Вкл' if notifications['shift1'] else 'Выкл'})",
                callback_data='toggle_shift1'
            ))
            markup.add(InlineKeyboardButton(
                f"2 Смена ({'Вкл' if notifications['shift2'] else 'Выкл'})",
                callback_data='toggle_shift2'
            ))
            markup.add(InlineKeyboardButton(
                f"3 Смена ({'Вкл' if notifications['shift3'] else 'Выкл'})",
                callback_data='toggle_shift3'
            ))
            markup.add(InlineKeyboardButton(
                f"Дежурство ({'Вкл' if notifications['weekend_duty'] else 'Выкл'})",
                callback_data='toggle_weekend_duty'
            ))
            markup.add(InlineKeyboardButton(
                f"Выходной ({'Вкл' if notifications['day_off'] else 'Выкл'})",
                callback_data='toggle_day_off'
            ))
            markup.add(InlineKeyboardButton(
                f"Время уведомлений: {notification_time}",
                callback_data='set_time'
            ))

            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text="Настройки уведомлений:",
                reply_markup=markup
            )
        elif call.data == 'set_time':
            msg = bot.send_message(call.message.chat.id,
                                   "Введите время для уведомлений в формате ЧЧ:ММ (например, 18:00)")
            bot.register_next_step_handler(msg, process_time_setting, login)


def process_time_setting(message, login):
    try:
        time_input = message.text.strip()
        # Проверка формата времени
        try:
            time_obj = datetime.strptime(time_input, "%H:%M")
            time_str = time_obj.strftime("%H:%M")
        except ValueError:
            bot.send_message(message.chat.id,
                             "Неверный формат времени. Используйте формат ЧЧ:ММ (например, 18:00)")
            settings_command(message)
            return

        user_data[login]["notification_time"] = time_str
        save_user_data(user_data)

        # Перепланируем все задачи
        schedule_updates()

        bot.send_message(message.chat.id, f"Время уведомлений установлено на {time_str}")
        settings_command(message)
    except Exception as e:
        logging.error(f"Error in process_time_setting: {e}")
        bot.send_message(message.chat.id, "Произошла ошибка при установке времени.")
        settings_command(message)



def get_weekday_name(date):
    weekdays = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']
    return weekdays[date.weekday()]

def get_shift_description(shift):
    if shift in [1, 2, 3]:
        return f"у вас {shift} смена."
    elif str(shift).upper() in ['В', 'B']:
        return "вы дежурите."
    elif shift is None or shift == '' or shift == 0:
        return "у вас выходной."
    else:
        return f"у вас смена: {shift}."


def calculate_worked_time(login, excel_file):
    print(f"Trying to open: {excel_file}")
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        current_date = datetime.now()
        first_day = current_date.replace(day=1)

        total_minutes = 0
        adjustment_minutes = 0
        working_days = 0  # Счетчик рабочих дней

        # Устанавливаем корректный индекс строки с датами
        date_row_index = 1  # Даты находятся в строке 1
        date_start_column = 4  # Даты начинаются с колонки 4

        # Создаем отображение даты на индекс колонки
        date_to_column = {}
        for col_idx, cell in enumerate(sheet[date_row_index][date_start_column - 1:], start=date_start_column):
            date_value = cell.value
            if isinstance(date_value, str):
                date_str = date_value.strip().lstrip('0')  # Убираем ведущие нули
                if date_str.isdigit():
                    date_to_column[date_str] = col_idx
            elif isinstance(date_value, int):
                date_str = str(date_value)
                date_to_column[date_str] = col_idx
            else:
                continue  # Пропускаем не-даты

        # Ищем строку с логином
        user_row = None
        user_row_idx = None
        for idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=False), start=3):
            cell_value = row[2].value  # Логин находится в колонке 3 (индекс 2)
            if cell_value:
                if str(cell_value).strip().lower() == login.lower():
                    user_row = row
                    user_row_idx = idx
                    break

        if not user_row:
            return "Логин не найден в расписании"

        # Собираем все корректировки времени из последней колонки
        comment_column_idx = sheet.max_column  # Предполагается, что комментарии в последней колонке
        comment_cell = sheet.cell(row=user_row_idx, column=comment_column_idx)

        # Парсим корректировки из комментария
        if comment_cell.value:
            adjustment_minutes += parse_time_adjustment(comment_cell.value)

        # Проходим по всем дням с начала месяца до текущей даты
        ru_holidays = holidays.RU()
        current_day = first_day
        while current_day <= current_date:
            weekday = current_day.weekday()
            date_str = current_day.strftime('%d').lstrip('0')  # Номер дня без ведущего нуля

            # Проверяем, является ли текущий день праздником-выходным
            if current_day in ru_holidays:
                # Отнимаем 30 минут за каждый праздник-выходной
                total_minutes -= 30
            # Получаем индекс колонки для текущей даты
            column_idx = date_to_column.get(date_str)
            if not column_idx:
                current_day += timedelta(days=1)
                continue

            # Получаем значение смены из ячейки
            cell = sheet.cell(row=user_row_idx, column=column_idx)
            shift_value = str(cell.value).strip() if cell.value else ''

            # Если смена указана (1, 2 или 3)
            if shift_value in ('1', '2', '3'):
                working_days += 1  # Увеличиваем счетчик рабочих дней

                # Определяем рабочее время для текущего дня
                if weekday == 4:  # Пятница
                    total_minutes += 6 * 60 + 55  # 6 часов 55 минут
                else:  # Пн-Чт
                    total_minutes += 7 * 60 + 25  # 7 часов 25 минут

            current_day += timedelta(days=1)

        # Конвертируем общее время в часы и минуты
        total_minutes += adjustment_minutes
        hours = total_minutes // 60
        minutes = total_minutes % 60

        # Формируем отчет
        report = f"Отработано с начала месяца по {current_date.strftime('%d.%m.%Y')}:\n\n"
        report += f"Рабочих дней: {working_days}\n"
        report += f"Всего: {hours} часов {minutes} минут\n"

        if adjustment_minutes != 0:
            adj_hours = abs(adjustment_minutes) // 60
            adj_minutes = abs(adjustment_minutes) % 60
            sign = '+' if adjustment_minutes > 0 else '-'
            report += f"\nКорректировки по комментариям: {sign}{adj_hours}ч {adj_minutes}м"

        return report

    except Exception as e:
        logging.error(f"Error calculating worked time: {e}")
        logging.exception("Full traceback:")
        return f"Ошибка при подсчете отработанного времени: {str(e)}"


def parse_adjustments_from_comment(comment):
    """Парсит корректировки времени из комментариев"""
    if not comment:
        return {}

    adjustments = {}

    # Ищем все даты и связанные с ними корректировки
    matches = re.findall(r'(\d{2}\.\d{2}\.\d{4})\s*([+-]?\d+ч\s*\d*м?)?', comment)
    for match in matches:
        date_str, time_str = match
        if time_str:
            minutes = parse_time_adjustment(time_str)
            adjustments[date_str] = adjustments.get(date_str, 0) + minutes

    return adjustments


def parse_time_adjustment(comment):
    """Парсит корректировки времени из комментария"""
    try:
        # Ищем все паттерны вида "+2ч 30м", "-35м", "-4ч 25м", "+1ч"
        total_minutes = 0

        # Ищем часы
        hours_matches = re.findall(r'([+-]?\d+)\s*ч', comment)
        for hours in hours_matches:
            total_minutes += int(hours) * 60

        # Ищем минуты
        minutes_matches = re.findall(r'([+-]?\d+)\s*м', comment)
        for minutes in minutes_matches:
            total_minutes += int(minutes)

        return total_minutes

    except Exception as e:
        logging.error(f"Error parsing time adjustment: {e}")
        return 0


def load_schedule_from_excel(filename):
    """
    Загружает расписание из указанного файла Excel.
    Возвращает список записей, где каждая запись — это словарь с данными пользователя.
    """
    try:
        logging.info(f"Загружаем расписание из файла: {filename}")

        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        schedule_data = []

        # Считываем данные из файла
        for row in sheet.iter_rows(min_row=4, values_only=True):  # Данные начинаются с 4-й строки
            logging.info(f"Обрабатываем строку: {row}")
            id = row[0]  # ID пользователя (1-я колонка)
            name = row[1]  # Имя пользователя (2-я колонка)
            login = row[2]  # Логин пользователя (3-я колонка)
            shifts = row[3:]  # Смены начинаются с 4-й колонки

            # Пропускаем строки без логина или имени
            if not login or not name:
                continue

            # Добавляем запись в расписание
            schedule_data.append({
                "id": id,
                "name": name.strip(),
                "login": login.strip(),
                "shifts": shifts
            })

        logging.info(f"Загружено записей из {filename}: {len(schedule_data)}")
        return schedule_data
    except Exception as e:
        logging.error(f"Ошибка при загрузке расписания из {filename}: {e}")
        return []


def get_user_info(login):
    """
    Получает информацию о сменах пользователя по логину.
    """
    logging.info(f"Ищем информацию о сменах для логина: {login}")

    # Загружаем данные из текущего расписания
    schedule_data = load_schedule_from_excel("schedule.xlsx")  # Текущее расписание
    next_schedule_data = load_schedule_from_excel("schedule_next.xlsx")  # Следующее расписание

    # Ищем пользователя в текущем расписании
    current_entry = next((entry for entry in schedule_data if entry["login"] == login), None)

    # Ищем пользователя в следующем расписании, если нужно
    next_entry = next((entry for entry in next_schedule_data if entry["login"] == login), None)

    if not current_entry and not next_entry:
        logging.error(f"Не найдена информация о сменах для логина: {login}")
        return None, None

    # Генерируем список смен
    shifts = []
    today = datetime.now().date()

    for i in range(4):  # Сегодня и следующие 3 дня
        check_date = today + timedelta(days=i)

        # Если дата в текущем месяце
        if check_date.month == today.month and current_entry:
            index = check_date.day - 1  # Индекс дня в массиве смен
            if index < len(current_entry["shifts"]):
                shifts.append((check_date, current_entry["shifts"][index]))
            else:
                shifts.append((check_date, None))

        # Если дата в следующем месяце
        elif next_entry and check_date.month == (today.month % 12) + 1:
            index = check_date.day - 1
            if index < len(next_entry["shifts"]):
                shifts.append((check_date, next_entry["shifts"][index]))
            else:
                shifts.append((check_date, None))

        else:
            # Если данные отсутствуют
            shifts.append((check_date, None))

    # Возвращаем список смен и имя пользователя
    name = current_entry.get("name") if current_entry else next_entry.get("name")
    return shifts, name

@bot.message_handler(commands=['shift'])
def shift_command(message):
    user_id = str(message.from_user.id)
    login = next((login for login, data in user_data.items() if data.get('user_id') == user_id), None)

    if login:
        # Получаем смены и имя сотрудника
        shifts, employee_name = get_user_info(login)

        if shifts:
            # Формируем ответ
            response = []
            for date, shift in shifts:
                date_str = date.strftime("%d.%m.%Y")
                weekday = get_weekday_name(date)
                shift_description = get_shift_description(shift)
                response.append(f"{date_str} ({weekday}) {shift_description}")

            # Отправляем результат
            bot.send_message(message.chat.id, "\n".join(response))
        else:
            bot.send_message(message.chat.id, "Не удалось найти информацию о ваших сменах.")
    else:
        bot.send_message(message.chat.id, "Вы не зарегистрированы в системе. Используйте /start для регистрации.")


@bot.message_handler(commands=['update'])
def update_command(message):
    user_id = str(message.from_user.id)
    logging.info(f"Update command received from user_id: {user_id}")

    if user_id in ADMIN_USERS:
        try:
            bot.reply_to(message, "Начинаю обновление расписания...")

            # Выполняем обновление расписания
            new_excel_file = download_and_process_doc()
            if new_excel_file:
                global EXCEL_FILE, excel_logins
                EXCEL_FILE = new_excel_file
                # Обновление данных
                excel_logins = load_excel_logins()
                schedule_common_jobs()
                schedule_updates()
                bot.reply_to(message, f"Расписание успешно обновлено!\nФайл: {new_excel_file}")
                logging.info(f"Schedule manually updated by admin {user_id}")
            else:
                bot.reply_to(message,
                             "Ошибка при обновлении расписания.\n"
                             "Проверьте логи и используйте /checkfile для диагностики.")
                logging.error(f"Manual schedule update failed by admin {user_id}")

        except Exception as e:
            error_message = f"Произошла ошибка при обновлении: {str(e)}"
            bot.reply_to(message, error_message)
            logging.error(f"Error during manual update by admin {user_id}: {e}")
            logging.exception("Full traceback:")
    else:
        bot.reply_to(message, "У вас нет прав для выполнения этой команды")

def kill_old_processes():
    """Завершение старых процессов бота"""
    current_pid = os.getpid()
    for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
        try:
            if proc.info['name'] == 'python.exe' and proc.pid != current_pid:
                cmdline = proc.info['cmdline']
                if cmdline and 'main.py' in cmdline[-1]:
                    proc.kill()
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            pass


@bot.message_handler(func=lambda message: True)
def handle_unknown(message):
    bot.reply_to(message, "Я не понимаю что вы пишете, это наверное что-то из 1С. Спросите у Н.Лагутини")

@tenacity.retry(wait=tenacity.wait_exponential(multiplier=1, min=4, max=10))
def start_bot():
    try:
        bot.polling(none_stop=True, interval=0, timeout=20)
    except Exception as e:
        logging.error(f"Ошибка опроса бота: {e}")
        raise


if __name__ == "__main__":
    while True:  # Бесконечный цикл
        try:
            fix_permissions()
            kill_old_processes()
            user_data = load_user_data()
            excel_logins = load_excel_logins()
            schedule_updates()

            scheduler_thread = Thread(target=run_scheduler, daemon=True)
            scheduler_thread.start()

            logging.info("Starting bot...")
            start_bot()

            # Бот запущен успешно, ждем прерывания
            # (например, Ctrl+C или сигнал завершения процесса)
            while True:
                time.sleep(60)  # проверяем каждую минуту, что бот жив. Можно увеличить интервал.
                if not scheduler_thread.is_alive():
                    logging.error("Scheduler thread died. Restarting bot...")
                    raise Exception("Scheduler thread died")  # вызываем исключение для перезапуска


        except Exception as e:
            logging.error(f"Critical error: {e}")
            logging.exception("Full traceback:")

            # Опционально: отправка сообщения об ошибке администратору
            # try:
            #   send_message_to_admin(f"Бот упал с ошибкой: {e}\nПолная трассировка в логах.")
            # except:
            #   logging.error("Не удалось отправить сообщение об ошибке админу.")

            logging.info("Restarting bot in 30 seconds...")
            time.sleep(30)  # Пауза перед перезапуском
            # После паузы цикл начнется снова и бот попытается перезапуститься

        finally:
            logging.info("Cleanup before restart (if needed)")
            # Очистка ресурсов при необходимости, например закрытие файлов, соединений с БД.
            # Это особенно важно, если у вас есть внешние ресурсы, которые нужно освободить.