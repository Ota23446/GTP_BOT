import asyncio
import json
import logging
import os
import pythoncom
import win32com.client
import openpyxl
from datetime import datetime, timedelta
from pathlib import Path
import requests
from config import MONTHS_RU, WEEKDAYS
from utils import parse_time
from aiogram.types import CallbackQuery

# Настройка логирования
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')


async def send_notifications(bot):
    """Общая функция для отправки уведомлений"""
    try:
        await asyncio.gather(
            send_shift_notifications(bot),
        )
    except Exception as e:
        logging.error(f"Error in send_notifications: {e}")


async def load_user_data():
    """Асинхронная загрузка данных пользователей"""
    try:
        return await asyncio.to_thread(_load_user_data_sync)
    except Exception as e:
        logging.error(f"Error loading user data: {e}")
        return {}


def _load_user_data_sync():
    """Синхронная загрузка данных пользователей"""
    try:
        with open('user_data.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        return {}


async def save_user_data(data):
    """Асинхронное сохранение данных пользователей"""
    try:
        await asyncio.to_thread(_save_user_data_sync, data)
    except Exception as e:
        logging.error(f"Error saving user data: {e}")


def _save_user_data_sync(data):
    """Синхронное сохранение данных пользователей"""
    with open('user_data.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4, ensure_ascii=False)


async def download_and_process_doc():
    """Асинхронная загрузка и обработка документа"""
    try:
        return await asyncio.to_thread(_download_and_process_doc_sync)
    except Exception as e:
        logging.error(f"Error in download_and_process_doc: {e}")
        return "Произошла ошибка при обработке документа"


def _download_and_process_doc_sync():
    """Синхронная загрузка и обработка документа"""
    url = "http://confluence.jira.lan:8090/exportword?pageId=24577712"
    doc_folder = "documents"
    excel_folder = "excel"
    main_schedule = 'schedule.xlsx'
    next_schedule = 'schedule_next.xlsx'

    # Создаем папки если их нет
    for folder in [doc_folder, excel_folder]:
        os.makedirs(folder, exist_ok=True)

    try:
        pythoncom.CoInitialize()

        # Скачивание файла
        response = requests.get(url, verify=False)
        response.raise_for_status()

        # Сохранение .doc файла
        current_date = datetime.now()
        doc_filename = f'schedule_{current_date.strftime("%Y%m%d")}.doc'
        doc_path = os.path.join(doc_folder, doc_filename)
        doc_path_abs = os.path.abspath(doc_path)

        with open(doc_path, 'wb') as f:
            f.write(response.content)

        # Обработка документа Word
        word = win32com.client.Dispatch('Word.Application')
        word.Visible = False

        try:
            doc = word.Documents.Open(doc_path_abs)

            # Получаем текущий и следующий месяц
            current_month = MONTHS_RU[current_date.month]
            current_year = current_date.strftime('%Y')
            current_header = f"{current_month} {current_year}"

            next_date = current_date + timedelta(days=32)
            next_month = MONTHS_RU[next_date.month]
            next_year = next_date.strftime('%Y')
            next_header = f"{next_month} {next_year}"

            # Создаем Excel файлы
            wb_current = openpyxl.Workbook()
            ws_current = wb_current.active
            wb_next = openpyxl.Workbook()
            ws_next = wb_next.active

            # Обработка таблиц
            found_current = False
            found_next = False

            for i in range(1, doc.Tables.Count + 1):
                table = doc.Tables.Item(i)

                # Получаем текст перед таблицей
                if i == 1:
                    range_start = 0
                else:
                    range_start = doc.Tables.Item(i - 1).Range.End
                range_end = table.Range.Start
                header_text = doc.Range(range_start, range_end).Text

                # Обработка таблицы текущего месяца
                if current_header in header_text:
                    found_current = True
                    _process_table(table, ws_current)

                # Обработка таблицы следующего месяца
                elif next_header in header_text:
                    found_next = True
                    _process_table(table, ws_next)

            # Сохранение файлов
            if found_current:
                wb_current.save(main_schedule)
            if found_next:
                wb_next.save(next_schedule)

            return f"Текущий месяц: {'обновлен' if found_current else 'не найден'}. Следующий месяц: {'обновлен' if found_next else 'не найден'}"

        finally:
            try:
                doc.Close(SaveChanges=False)
                word.Quit()
            except:
                pass

    except Exception as e:
        logging.error(f"Error processing document: {e}")
        return f"Ошибка обработки документа: {str(e)}"
    finally:
        pythoncom.CoUninitialize()


def _process_table(table, worksheet):
    """Обработка отдельной таблицы"""
    rows = table.Rows.Count
    cols = table.Columns.Count

    for row in range(1, rows + 1):
        for col in range(1, cols + 1):
            try:
                cell_value = table.Cell(Row=row, Column=col).Range.Text
                cell_value = cell_value.strip().replace('\r', '').replace('\a', '')
                worksheet.cell(row=row, column=col, value=cell_value)
            except Exception as e:
                logging.warning(f"Error processing cell [{row},{col}]: {e}")


async def calculate_worked_time(login):
    """Асинхронный подсчет отработанного времени"""
    try:
        return await asyncio.to_thread(_calculate_worked_time_sync, login)
    except Exception as e:
        logging.error(f"Error calculating worked time: {e}")
        return "Ошибка при подсчете отработанного времени"


def _calculate_worked_time_sync(login):
    """Синхронный подсчет отработанного времени"""
    try:
        wb = openpyxl.load_workbook('schedule.xlsx')
        ws = wb.active

        # Поиск логина в таблице
        login_cell = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).lower() == login.lower():
                    login_cell = cell
                    break
            if login_cell:
                break

        if not login_cell:
            return "Логин не найден в расписании"

        # Подсчет отработанных смен
        row = login_cell.row
        total_hours = 0
        shifts = {
            "1": 11,  # 1 смена - 11 часов
            "2": 11,  # 2 смена - 11 часов
            "3": 11,  # 3 смена - 11 часов
            "В": 12,  # Выходной день (дежурство) - 12 часов
            "B": 12  # Учитываем английскую B для случая опечаток
        }

        # Получаем текущий день месяца
        current_day = datetime.now().day

        # Подсчет часов для каждого дня до текущего
        for col in range(2, current_day + 2):  # +2 так как первая колонка - это логин
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                cell_value = str(cell_value).strip()
                hours = shifts.get(cell_value, 0)
                total_hours += hours

        # Форматирование результата
        days = total_hours // 24
        remaining_hours = total_hours % 24

        result = f"Отработано с начала месяца:\n"
        result += f"Всего часов: {total_hours}\n"
        if days > 0:
            result += f"Это составляет {days} дней и {remaining_hours} часов"
        else:
            result += f"Это составляет {remaining_hours} часов"

        return result

    except Exception as e:
        logging.error(f"Error in calculate_worked_time: {e}")
        return "Ошибка при подсчете отработанного времени"
    finally:
        wb.close()


async def send_shift_notifications(bot):
    """Отправка уведомлений о сменах"""
    try:
        user_data = await load_user_data()  # Предполагается, что эта функция есть
        current_date = datetime.now()
        tomorrow = current_date + timedelta(days=1)

        wb = await asyncio.to_thread(openpyxl.load_workbook, 'schedule.xlsx')
        ws = wb.active

        login_row_map = {}
        for row in ws.iter_rows():
            login = str(row[2].value).lower() if row[2].value else None
            if login:
                login_row_map[login] = row[2].row

        for login, data in user_data.items():
            try:
                user_id = data.get('user_id')
                if not user_id:
                    logging.warning(f"No user_id found for {login}")
                    continue

                notifications = data.get('notifications', {})
                notification_time = data.get('notification_time', "18:00")

                current_time = parse_time(current_date.strftime("%H:%M"))
                notification_time_obj = parse_time(notification_time)

                if current_time != notification_time_obj:
                    continue


                login_row = login_row_map.get(login.lower())
                if login_row:
                    tomorrow_col = tomorrow.day + 3  #
                    shift = ws.cell(row=login_row, column=tomorrow_col).value

                    if shift:
                        shift = str(shift).strip()
                        message = None

                        if shift == "1" and notifications.get('shift1'):
                            message = f"Завтра {tomorrow.strftime('%d.%m.%Y')} у вас 1 смена"
                        elif shift == "2" and notifications.get('shift2'):
                            message = f"Завтра {tomorrow.strftime('%d.%m.%Y')} у вас 2 смена"
                        elif shift == "3" and notifications.get('shift3'):
                            message = f"Завтра {tomorrow.strftime('%d.%m.%Y')} у вас 3 смена"
                        elif shift.upper() in ['В', 'B'] and notifications.get('weekend'):
                            message = f"Завтра {tomorrow.strftime('%d.%m.%Y')} у вас дежурство"
                        elif notifications.get('dayoff'):
                            message = f"Завтра {tomorrow.strftime('%d.%m.%Y')} у вас выходной"

                        if message:
                            await bot.send_message(user_id, message)

            except Exception as e:
                logging.error(f"Error sending notification to {login}: {e}")

        wb.close()

    except Exception as e:
        logging.error(f"Error in send_shift_notifications: {e}")


async def start_notification_scheduler(bot):
    """Запуск планировщика уведомлений"""
    while True:
        try:
            await send_shift_notifications(bot)
            await asyncio.sleep(60)  # Проверка каждую минуту
        except Exception as e:
            logging.error(f"Error in notification scheduler: {e}")
            await asyncio.sleep(60)


async def send_monday_notification(bot):
    """Отправка уведомлений по понедельникам"""
    user_data = await load_user_data()
    for _, data in user_data.items():
        try:
            user_id = data.get('user_id')
            await bot.send_message(user_id, "Завтра вторник, нужно проверить битрикс на просроченные задачи.")
        except Exception as e:
            logging.error(f"Error sending Monday notification: {e}")


async def send_tuesday_notification(bot):
    """Отправка уведомлений по вторникам"""
    user_data = await load_user_data()
    for _, data in user_data.items():
        try:
            user_id = data.get('user_id')
            await bot.send_message(user_id, "Сегодня вторник, нужно проверить битрикс на просроченные задачи.")
        except Exception as e:
            logging.error(f"Error sending Tuesday notification: {e}")



# Добавляем вспомогательную функцию для проверки активности пользователей
async def check_user_activity(bot, days_inactive: int = 30):
    """Проверка активности пользователей"""
    try:
        user_data = await load_user_data()
        inactive_users = []

        for login, data in user_data.items():
            try:
                user_id = data.get('user_id')
                chat = await bot.get_chat(user_id)
                # Здесь можно добавить дополнительную логику проверки активности
            except Exception as e:
                inactive_users.append(login)
                logging.error(f"Error checking activity for {login}: {e}")

        return inactive_users
    except Exception as e:
        logging.error(f"Error in check_user_activity: {e}")
        return []


async def get_next_shift(login):
    """Получение информации о следующих сменах"""
    try:
        logging.info(f"Getting next shifts for user {login}")

        # Получаем даты для трех следующих дней
        current_date = datetime.now()
        next_dates = [
            current_date + timedelta(days=i + 1)
            for i in range(3)
        ]

        shifts_info = []
        current_month = current_date.month

        for date in next_dates:
            # Определяем, какой файл использовать
            if date.month == current_month:
                filename = 'schedule.xlsx'
            else:
                filename = 'schedule_next.xlsx'

            try:
                wb = openpyxl.load_workbook(filename)
                ws = wb.active

                # Поиск логина в таблице (в 3-й колонке)
                login_cell = None
                login_row = None
                for row in ws.iter_rows():
                    cell = row[2]  # 3-я колонка (индекс 2)
                    cell_value = str(cell.value).lower() if cell.value else ''
                    if cell_value == login.lower():
                        login_cell = cell
                        login_row = cell.row
                        break

                if not login_cell:
                    shifts_info.append(
                        f"{date.strftime('%d.%m.%Y')} ({WEEKDAYS[date.strftime('%A').lower()]}) - логин не найден")
                    continue

                # Смены начинаются с 4-й колонки, поэтому добавляем 3 к номеру дня
                day_col = date.day + 2
                shift = ws.cell(row=login_row, column=day_col).value

                # Получаем день недели на русском
                weekday = WEEKDAYS[date.strftime('%A').lower()]

                if shift:
                    shift = str(shift).strip()
                    is_friday = date.weekday() == 4

                    if shift == "1":
                        if is_friday:
                            time_info = "08:00 - 16:00"
                        else:
                            time_info = "08:00 - 16:30"
                        shift_text = f"первая смена ({time_info})"

                    elif shift == "2":
                        if is_friday:
                            time_info = "09:30 - 17:30"
                        else:
                            time_info = "09:30 - 18:00"
                        shift_text = f"вторая смена ({time_info})"

                    elif shift == "3":
                        if is_friday:
                            time_info = "12:00 - 20:00"
                        else:
                            time_info = "11:30 - 20:00"
                        shift_text = f"третья смена ({time_info})"

                    elif shift.upper() in ['В', 'B', 'в']:
                        shift_text = "дежурство в выходной день"

                    else:
                        shift_text = "выходной"

                    shifts_info.append(f"{date.strftime('%d.%m.%Y')} ({weekday}) - {shift_text}")
                else:
                    shifts_info.append(f"{date.strftime('%d.%m.%Y')} ({weekday}) - смена не назначена")

                wb.close()

            except FileNotFoundError:
                shifts_info.append(
                    f"{date.strftime('%d.%m.%Y')} ({WEEKDAYS[date.strftime('%A').lower()]}) - расписание не найдено")
                logging.error(f"Schedule file not found: {filename}")
            except Exception as e:
                shifts_info.append(
                    f"{date.strftime('%d.%m.%Y')} ({WEEKDAYS[date.strftime('%A').lower()]}) - ошибка чтения")
                logging.error(f"Error reading schedule for date {date}: {e}")

        response = "📅 Расписание на ближайшие дни:\n\n"
        response += "\n".join(shifts_info)

        logging.info(f"Prepared response for {login}: {response}")
        return response

    except Exception as e:
        logging.error(f"Error in get_next_shift for {login}: {e}")
        return f"Ошибка при получении информации о сменах: {str(e)}"

async def get_shift_for_date(login, target_date, filename):
    """Получает смену пользователя на заданную дату из указанного файла"""
    try:
        wb = await asyncio.to_thread(openpyxl.load_workbook, filename)
        ws = wb.active

        login_row_map = {}
        for row in ws.iter_rows():
            login_val = str(row[2].value).lower() if row[2].value else None
            if login_val:
                login_row_map[login_val] = row[2].row

        login_row = login_row_map.get(login.lower())
        if not login_row:
            return f"Логин {login} не найден в расписании."

        day_col = target_date.day + 3
        shift_value = ws.cell(row=login_row, column=day_col).value
        wb.close()

        if shift_value:
            shift_str = str(shift_value).strip()
            if shift_str == "1":
                shift_text = "первая смена"
            elif shift_str == "2":
                shift_text = "вторая смена"
            elif shift_str == "3":
                shift_text = "третья смена"
            elif shift_str.upper() in ['В', 'B', 'в']:
                shift_text = "дежурство"
            else:
                shift_text = "выходной"

            return f"{target_date.strftime('%d.%m.%Y')} - {shift_text}"

        else:
            return f"Смена не назначена на {target_date.strftime('%d.%m.%Y')}"

    except FileNotFoundError:
        return "Файл расписания не найден."
    except Exception as e:
        logging.error(f"Error getting shift for date: {e}")
        return f"Ошибка при получении смены: {e}"

async def process_schedule_day(callback: CallbackQuery, month: int, day: int):
    """Обрабатывает выбор конкретной даты и показывает смену"""
    user_id = str(callback.from_user.id)
    username = user_manager.get_user_by_telegram_id(user_id)
    if not username:
        await callback.message.answer("Сначала зарегистрируйтесь! /register")
        return

    current_date = datetime.now()
    year = current_date.year
    if month > 12:
        year += 1

    target_date = datetime(year, month, day)

    filename = "schedule.xlsx" if month == current_date.month else "schedule_next.xlsx"
    shift_info = await get_shift_for_date(username, target_date, filename)
    await callback.message.answer(shift_info)